import os
from openpyxl import load_workbook

# Get the directory of the current script
script_dir = os.path.dirname(os.path.abspath(__file__))

# Construct the path to the Excel template using a relative path
template_path = os.path.join(script_dir, 'Invoice temple_CCC.xlsx')
formatted_data_2 = ''

def create_excel_invoice(invoice_data):
    global formatted_data_2        
    # template_path = r"C:\Users\13473\Desktop\CCC\Inovices\Invoice temple_CCC.xlsx"  # Update this path to your template's location
    invoices_folder = os.path.expanduser("~/Desktop/invoices")
    if not os.path.exists(invoices_folder):
        os.makedirs(invoices_folder)

    # Load the workbook and select the active worksheet
    wb = load_workbook(template_path)
    ws = wb.active

    # Fill in the details in the Excel sheet
    ws['H4'] = ws['C13'] = ws['D13'] = invoice_data['INV#']
    ws['C6'] = ws['F6'] = invoice_data['SOLD TO']
    ws['H3'] = invoice_data['DATE']
    ws['C7'] = ws['F7'] = invoice_data['Address']
    ws['B13'] = invoice_data['SALESMAN']
    ws['C8'] = invoice_data['City']
    ws['F8'] = invoice_data['City']
    ws['D8'] = invoice_data['State'] + ' ' + invoice_data['ZIP']  # Assuming State and ZIP should be in the same cell separated by a space
    ws['G8'] = invoice_data['State'] + ' ' + invoice_data['ZIP'] 
    ws['E13'] = invoice_data['Shipping Terms']
    ws['F13'] = invoice_data['Amount down']
    ws['G13'] = invoice_data['Payment Method']
    ws['H13'] = invoice_data['Due Date']

    # Fill in the line items
    line_item_fields = ['QTY', 'DATE(Year)', 'GRADE', 'Description', 'Unit Price']
    line_item_cells = {
        'QTY': ['B17', 'B19', 'B21', 'B23', 'B25'],
        'DATE(Year)': ['C17', 'C19', 'C21', 'C23', 'C25'],
        'GRADE': ['D17', 'D19', 'D21', 'D23', 'D25'],
        'Description': ['E17', 'E19', 'E21', 'E23', 'E25'],
        'Unit Price': ['F17', 'F19', 'F21', 'F23', 'F25']
    }

    # Before processing line items, find the minimum length of the provided lists to avoid "index out of range" errors
    min_line_items_length = min(len(invoice_data[field]) for field in line_item_fields)

    for field in line_item_fields:
        for i in range(min_line_items_length):
            value = invoice_data[field][i]
            cell = line_item_cells[field][i]
            if value:  # Check if there's actually a value to avoid writing empty strings for missing data
                ws[cell] = value


    # Calculate Line Totals and Subtotal
    # Calculate Line Totals and Subtotal
    subtotal = 0
    for i, qty_str in enumerate(invoice_data['QTY']):
        # Check if qty_str is not empty, else default to 0
        qty = int(qty_str) if qty_str else 0
        
        unit_price_str = invoice_data['Unit Price'][i]
        # Check if unit_price_str is not empty, else default to 0.0
        unit_price = float(unit_price_str) if unit_price_str else 0.0
        
        line_total = qty * unit_price
        if line_total != 0:  # Only write line_total if it's not 0
            ws[line_item_cells['QTY'][i].replace('B', 'H')] = line_total
        subtotal += line_total

    # Write Subtotal, Shipping, and Balance Due only if they are not 0
    if subtotal != 0:
        ws['H38'] = subtotal

    try:
        shipping_str = invoice_data.get('Shipping', '0')  # Default to '0' if not provided
        shipping_cost = float(shipping_str) if shipping_str else 0.0
    except ValueError:
        shipping_cost = 0.0  # Default to 0.0 if conversion fails
    try:
        amount_down_str = invoice_data.get('Amount down', '0')  # Default to '0' if not provided
        amount_down = float(amount_down_str) if amount_down_str else 0.0
    except ValueError:
        amount_down = 0.0  # Default to 0.0 if conversion fails
    if shipping_cost != 0.0:
        ws['H39'] = shipping_cost

    balance_due = subtotal - amount_down
    if balance_due != 0.0:
        ws['H40'] = balance_due

    costumer_name = invoice_data['SOLD TO']
    first_initial = costumer_name.split()[0][0]
    last_name = costumer_name.split()[-1]
    shortened_name = f"{first_initial}. {last_name}"
    # Save the filled-in invoice to a new file
    new_invoice_path = os.path.join(invoices_folder, f"{shortened_name} Inv {invoice_data['INV#']}.xlsx")
    wb.save(new_invoice_path)
    
    # Close the workbook
    wb.close()
    details = [f"({qty}) {grade} {desc}" for qty, grade, desc in zip(invoice_data['QTY'], invoice_data['GRADE'], invoice_data['Description'])]
    combined_string = ' (), '.join(details)
    formatted_data_2 = f"{invoice_data['DATE']}\t{invoice_data['INV#']}\t{shortened_name}\t{invoice_data['SALESMAN']}\t\"{combined_string}\""
    print(formatted_data_2)

    return new_invoice_path, formatted_data_2  # Returns the path and data for global use



# Example usage (you need to replace 'path_to_template/template.xlsx' with your actual template's path and provide the actual data in 'invoice_data'):
# invoice_data = {
#     'INV#': '12345',
#     ... # other fields and values as required
# }
# new_invoice_path = create_excel_invoice(invoice_data)
# print("Invoice saved to:", new_invoice_path)