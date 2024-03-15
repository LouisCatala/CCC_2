# Make a better code, consider goal, the process,
# after (the possible changes required), Connect with acutal situation. 
# pyinstaller --onefile --windowed --add-data "Invoice temple_CCC.xlsx;." -n Create_Invoice Section1.py
import os
from openpyxl import load_workbook
import subprocess
import sys


# Get the directory of the current script
script_dir = os.path.dirname(os.path.abspath(__file__))

# Construct the path to the Excel template using a relative path
template_path = os.path.join(script_dir, 'Invoice temple_CCC.xlsx')

#to Open Excel after run
def open_file(file_path):
    if sys.platform == "win32":
        os.startfile(file_path)
    elif sys.platform == "darwin":
        subprocess.Popen(["open", file_path])
    else:  # 'linux' or 'linux2' for Linux
        subprocess.Popen(["xdg-open", file_path])
        
def create_excel_invoice(invoice_data):
    global formatted_data_2        
    # template_path = r"C:\Users\13473\Desktop\CCC\Inovices\Invoice temple_CCC.xlsx"  # Update this path to your template's location
    invoices_folder = os.path.expanduser("~/Desktop/invoices")
    if not os.path.exists(invoices_folder):
        os.makedirs(invoices_folder)

    # Load the workbook and select the active worksheet
    wb = load_workbook(template_path)
    ws = wb.active

    # Fill in the details in the Excel sheet
    ws['H4'] = ws['C13'] = ws['D13'] = invoice_data['INV#']
    ws['C6'] = ws['F6'] = invoice_data['SOLD TO']
    ws['H3'] = invoice_data['DATE']
    ws['C7'] = ws['F7'] = invoice_data['Address']
    ws['B13'] = invoice_data['SALESMAN']
    ws['C8'] = invoice_data['City']
    ws['F8'] = invoice_data['City']
    ws['D8'] = invoice_data['State'] + ' ' + invoice_data['ZIP']  # Assuming State and ZIP should be in the same cell separated by a space
    ws['G8'] = invoice_data['State'] + ' ' + invoice_data['ZIP'] 
    ws['E13'] = invoice_data['Shipping Terms']
    ws['F13'] = invoice_data['Amount down']
    ws['G13'] = invoice_data['Payment Method']
    ws['H13'] = invoice_data['Due Date']

    # Fill in the line items
    line_item_fields = ['QTY', 'DATE(Year)', 'GRADE', 'Description', 'Unit Price']
    line_item_cells = {
        'QTY': ['B17', 'B19', 'B21', 'B23', 'B25', 'B27', 'B29', 'B31', 'B33', 'B35'],
        'DATE(Year)': ['C17', 'C19', 'C21', 'C23', 'C25', 'C27', 'C29', 'C31', 'C33', 'C35'],
        'GRADE': ['D17', 'D19', 'D21', 'D23', 'D25', 'D27', 'D29', 'D31', 'D33', 'D35'],
        'Description': ['E17', 'E19', 'E21', 'E23', 'E25', 'E27', 'E29', 'E31', 'E33', 'E35'],
        'Unit Price': ['F17', 'F19', 'F21', 'F23', 'F25', 'F27', 'F29', 'F31', 'F33', 'F35']
    }


    # Before processing line items, find the minimum length of the provided lists to avoid "index out of range" errors
    min_line_items_length = min(len(invoice_data[field]) for field in line_item_fields)

    for field in line_item_fields:
        for i in range(min_line_items_length):
            value = invoice_data[field][i]
            cell = line_item_cells[field][i]
            if value:  # Check if there's actually a value to avoid writing empty strings for missing data
                ws[cell] = value


    # Calculate Line Totals and Subtotal
    # Calculate Line Totals and Subtotal
    subtotal = 0
    for i, qty_str in enumerate(invoice_data['QTY']):
        # Check if qty_str is not empty, else default to 0
        qty = int(qty_str) if qty_str else 0
        
        unit_price_str = invoice_data['Unit Price'][i]
        # Check if unit_price_str is not empty, else default to 0.0
        unit_price = float(unit_price_str) if unit_price_str else 0.0
        
        line_total = qty * unit_price
        if line_total != 0:  # Only write line_total if it's not 0
            ws[line_item_cells['QTY'][i].replace('B', 'H')] = line_total
        subtotal += line_total

    # Write Subtotal, Shipping, and Balance Due only if they are not 0
    if subtotal != 0:
        ws['H38'] = subtotal

    try:
        shipping_str = invoice_data.get('Shipping', '0')  # Default to '0' if not provided
        shipping_cost = float(shipping_str) if shipping_str else 0.0
    except ValueError:
        shipping_cost = 0.0  # Default to 0.0 if conversion fails
    try:
        amount_down_str = invoice_data.get('Amount down', '0')  # Default to '0' if not provided
        amount_down = float(amount_down_str) if amount_down_str else 0.0
    except ValueError:
        amount_down = 0.0  # Default to 0.0 if conversion fails
    if shipping_cost != 0.0:
        ws['H39'] = shipping_cost

    balance_due = subtotal - amount_down
    if balance_due != 0.0:
        ws['H40'] = balance_due

    costumer_name = invoice_data['SOLD TO']
    first_initial = costumer_name.split()[0][0]
    last_name = costumer_name.split()[-1]
    shortened_name = f"{first_initial}. {last_name}"
    # Save the filled-in invoice to a new file
    new_invoice_path = os.path.join(invoices_folder, f"{shortened_name} Inv {invoice_data['INV#']}.xlsx")
    wb.save(new_invoice_path)
    
    # Close the workbook
    wb.close()
    open_file(new_invoice_path) #Open the Excel
    details = [f"({qty}) {grade} {desc}" for qty, grade, desc in zip(invoice_data['QTY'], invoice_data['GRADE'], invoice_data['Description'])]
    combined_string = ' (), '.join(details)
    formatted_data_2 = f"{invoice_data['DATE']}\t{invoice_data['INV#']}\t{shortened_name}\t{invoice_data['SALESMAN']}\t\"{combined_string}\"" 
    return new_invoice_path, formatted_data_2 


import tkinter as tk
from tkinter import messagebox  


def focus_next_widget(event):
    event.widget.tk_focusNext().focus()
    return("break")
    
app = tk.Tk()
app.title("Invoice Input Form")

main_frame = tk.Frame(app)
main_frame.pack(pady=10, padx=10, fill="both", expand=True)

instruction_label = tk.Label(main_frame, text="Press Tab/ enter go Next box. After click on text box, \n if needed, Ctrl+A to select infos. Ctrl+C to copy, Ctrl+V to paste infos")
instruction_label.grid(row=0, column=0, columnspan=2, pady=5)

# Define the regular fields
regular_fields = [
    'INV#', 'SOLD TO', 'DATE', 'Address', 'SALESMAN', 
    'City', 'State', 'ZIP', 'Shipping Terms', 'Amount down', 
    'Payment Method', 'Due Date'
]

entries = {}
# Create a frame for the regular fields
left_frame = tk.Frame(main_frame)
left_frame.grid(row=1, column=0, sticky="ns")

# Add fields to the left_frame
for idx, field in enumerate(regular_fields):
    label = tk.Label(left_frame, width=15, text=field+": ", anchor='w')
    entry = tk.Entry(left_frame)
    entry.bind("<Return>", focus_next_widget)
    label.grid(row=idx+1, column=0, sticky="e")
    entry.grid(row=idx+1, column=1, sticky="ew")
    entries[field] = entry

# Handling multiple entries, create a frame for the multi-entry fields
right_frame = tk.Frame(main_frame)
right_frame.grid(row=1, column=1, sticky="ns")

multi_entry_fields = [
    ('QTY', 10),
    ('DATE(Year)', 10),
    ('GRADE', 10),
    ('Description', 10),
    ('Unit Price', 10)
]

for idx, (field, num) in enumerate(multi_entry_fields):
    label = tk.Label(right_frame, text=field+":", anchor='e')
    label.grid(row=idx, column=0, sticky='e')
    entries[field] = []
    for i in range(num):
        entry = tk.Entry(right_frame)
        entry.bind("<Return>", focus_next_widget)
        # Apply internal padding to make the entry widget appear taller
        entry.grid(row=idx, column=i+1, padx=2, sticky="ew", ipady=15)  # Adjust ipady as needed
        entries[field].append(entry)


# Ensure the grid cells in the left frame expand as needed
left_frame.grid_columnconfigure(1, weight=1)

# Ensure the grid cells in the right frame expand as needed
for i in range(1, 11):  # Now configuring for 10 entries
    right_frame.grid_columnconfigure(i, weight=1)

# Function for the Submit button
def on_submit():
    global global_invoice_data
    invoice_data = {}
    for field, entry in entries.items():
        if isinstance(entry, list):  # For fields with multiple entries
            invoice_data[field] = [e.get() for e in entry]
        else:
            invoice_data[field] = entry.get()
    global_invoice_data = invoice_data
    
    # Assuming create_excel_invoice is properly imported or defined in this script
    try:
        new_invoice_path, formatted_data_2 = create_excel_invoice(invoice_data)
        print(f"Invoice saved to: {new_invoice_path}")
        messagebox.showinfo("Success", f"Invoice saved to: {new_invoice_path}")
    except Exception as e:
        messagebox.showerror("Error", f"Error generating invoice: {e}")
        print(f"Error generating invoice: {e}")
        
# Copy Button for copy and paste
def copy_to_clipboard():
    app.clipboard_clear()  # Clear the clipboard
    app.clipboard_append(formatted_data_2)  # Append the formatted data to the clipboard
    print("Data copied to clipboard.")

copy_button = tk.Button(app, text='Copy', command=copy_to_clipboard)
copy_button.pack(side=tk.RIGHT, padx=0, pady=5)

# Submit button
submit_button = tk.Button(app, text='Submit', command=on_submit)
submit_button.pack(side=tk.RIGHT, padx=5, pady=5)


if __name__ == "__main__":
    app.mainloop()